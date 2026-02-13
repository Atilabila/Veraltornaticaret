# -*- coding: utf-8 -*-
"""
Shared Excel template generator for /admin/urunler bulk upload.

This module is imported by both:
- scripts/generate_bulk_urunler_template.py (CLI)
- scripts/bulk_excel_wizard.py (GUI)

Keeping the logic here makes PyInstaller packaging reliable.
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "Ürün İsmi*",
    "Fiyat*",
    "Görsel URL*",
    "Açıklama",
    "Özellik 1",
    "Özellik 2",
    "Özellik 3",
    "Özellik 4",
]

# Column widths tuned for copy/paste workflow.
COLUMN_WIDTHS = [32, 14, 56, 44, 26, 26, 26, 26]


@dataclass(frozen=True)
class Theme:
    header_fill: PatternFill
    header_font: Font
    header_alignment: Alignment
    required_bad_fill: PatternFill
    warning_fill: PatternFill
    body_alignment: Alignment


THEME = Theme(
    header_fill=PatternFill("solid", fgColor="0F172A"),  # slate-950
    header_font=Font(bold=True, color="FFFFFF"),
    header_alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
    required_bad_fill=PatternFill("solid", fgColor="FEE2E2"),  # red-100
    warning_fill=PatternFill("solid", fgColor="FEF3C7"),  # amber-100
    body_alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
)


def _apply_header_styles(ws) -> None:
    ws.append(HEADERS)
    ws.row_dimensions[1].height = 28

    for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=HEADERS[col_idx - 1])
        cell.fill = THEME.header_fill
        cell.font = THEME.header_font
        cell.alignment = THEME.header_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


def _apply_body_styles(ws, max_rows: int) -> None:
    for row in range(2, max_rows + 2):
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row, column=col).alignment = THEME.body_alignment

    url_font = Font(size=10)
    for row in range(2, max_rows + 2):
        ws.cell(row=row, column=3).font = url_font


def _apply_conditional_formatting(ws, max_rows: int) -> None:
    last_row = max_rows + 1

    # Required fields: Name, Price, Image URL.
    required_ranges = {
        "A": f"A2:A{last_row}",
        "B": f"B2:B{last_row}",
        "C": f"C2:C{last_row}",
    }

    ws.conditional_formatting.add(
        required_ranges["A"],
        FormulaRule(formula=['LEN(TRIM($A2))=0'], stopIfTrue=False, fill=THEME.required_bad_fill),
    )
    ws.conditional_formatting.add(
        required_ranges["B"],
        FormulaRule(formula=['LEN(TRIM($B2&""))=0'], stopIfTrue=False, fill=THEME.required_bad_fill),
    )
    ws.conditional_formatting.add(
        required_ranges["C"],
        FormulaRule(formula=['LEN(TRIM($C2))=0'], stopIfTrue=False, fill=THEME.required_bad_fill),
    )

    # URL format warnings (http/https).
    ws.conditional_formatting.add(
        required_ranges["C"],
        FormulaRule(
            formula=[
                'AND(LEN(TRIM($C2))>0,NOT(OR(LEFT($C2,7)="http://",LEFT($C2,8)="https://")))'
            ],
            stopIfTrue=False,
            fill=THEME.warning_fill,
        ),
    )

    # Duplicate name warning within the sheet.
    ws.conditional_formatting.add(
        required_ranges["A"],
        FormulaRule(
            formula=['AND(LEN(TRIM($A2))>0,COUNTIF($A:$A,$A2)>1)'],
            stopIfTrue=False,
            fill=THEME.warning_fill,
        ),
    )


def _seed_example_rows(ws) -> None:
    ws.append(
        [
            "Metal Poster Örnek 01",
            "350",
            "https://example.com/images/metal-poster-01.webp",
            "Kısa açıklama (opsiyonel).",
            "1.5mm metal",
            "UV baskı",
            "Hazır askı",
            "Yerli üretim",
        ]
    )
    ws.append(
        [
            "Metal Poster Örnek 02",
            "₺499,90",
            "https://example.com/images/metal-poster-02.webp",
            "",
            "Mat boya",
            "",
            "",
            "",
        ]
    )


def build_workbook(max_rows: int, include_examples: bool) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "URUNLER"

    _apply_header_styles(ws)
    if include_examples:
        _seed_example_rows(ws)

    _apply_body_styles(ws, max_rows=max_rows)
    _apply_conditional_formatting(ws, max_rows=max_rows)

    # Add short usage note in the sheet (kept out of the 8 columns area).
    note_col = get_column_letter(len(HEADERS) + 2)
    ws[f"{note_col}1"] = "KULLANIM"
    ws[f"{note_col}1"].font = Font(bold=True)
    ws[f"{note_col}2"] = "1) Excel'de 8 kolonu (A-H) doldur."
    ws[f"{note_col}3"] = "2) /admin/urunler sayfasında kategori seç."
    ws[f"{note_col}4"] = "3) A-H aralığını (başlık hariç) kopyala-yapıştır, Toplu Kaydet."
    ws.column_dimensions[note_col].width = 46

    return wb

